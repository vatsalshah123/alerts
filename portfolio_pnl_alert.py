"""
Portfolio P&L Alert — Telegram + Email (both free)
-----------------------------------------------------
Reads holdings from portfolio.xlsx, fetches live NSE/BSE prices,
computes P&L, and sends the summary to BOTH Telegram (headline summary
message + a PDF attachment with the full per-stock table) and Email
(real HTML table via Brevo, works great in Hotmail/Outlook/Gmail).

SETUP:
1. pip install -r requirements.txt
2. Telegram: create a bot via @BotFather, get BOT_TOKEN + CHAT_ID (see README)
3. Email: create a free Brevo account (brevo.com), verify a sender email,
   get an API key. No App Passwords / 2FA hassle needed. (see README)
4. Fill in credentials below (or set as env vars / GitHub Secrets)
5. Test: python portfolio_pnl_alert.py
6. Schedule hourly (see README)

portfolio.xlsx format (first sheet, header row required):
| symbol   | exchange | qty | buy_price |
|----------|----------|-----|-----------|
| RELIANCE | NSE      | 10  | 2450.50   |
| TCS      | BSE      | 5   | 3600.00   |
"""

import os
import sys
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from io import BytesIO
from zoneinfo import ZoneInfo

import requests
import yfinance as yf
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

# ---------------- CONFIG ----------------
# All read from environment variables (set as GitHub Secrets in Actions).
BOT_TOKEN = os.environ.get("BOT_TOKEN")
CHAT_ID = os.environ.get("CHAT_ID")
BREVO_API_KEY = os.environ.get("BREVO_API_KEY")
EMAIL_FROM = os.environ.get("EMAIL_FROM")
EMAIL_TO = os.environ.get("EMAIL_TO")

PORTFOLIO_FILE = "portfolio.xlsx"
# ------------------------------------------

EXCHANGE_SUFFIX = {"NSE": ".NS", "BSE": ".BO"}

IST = ZoneInfo("Asia/Kolkata")


def now_ist():
    """GitHub Actions runners run in UTC — always report timestamps in IST."""
    return datetime.now(IST)


def load_portfolio(path):
    """Reads holdings from the first sheet of an .xlsx file.
    Expects header row: symbol | exchange | qty | buy_price"""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    col = {name: idx for idx, name in enumerate(header)}

    required = {"symbol", "exchange", "qty", "buy_price"}
    missing = required - set(col)
    if missing:
        raise ValueError(f"portfolio.xlsx is missing column(s): {missing}")

    holdings = []
    for r in rows[1:]:
        if not r or r[col["symbol"]] in (None, ""):
            continue  # skip blank rows
        holdings.append({
            "symbol": str(r[col["symbol"]]).strip(),
            "exchange": str(r[col["exchange"]]).strip(),
            "qty": float(r[col["qty"]]),
            "buy_price": float(r[col["buy_price"]]),
        })
    return merge_duplicate_holdings(holdings)


def merge_duplicate_holdings(holdings):
    """Combines multiple rows for the same stock (symbol + exchange) into
    one, using a quantity-weighted average buy price."""
    merged = {}
    order = []  # preserves first-seen order for consistent report ordering

    for h in holdings:
        key = (h["symbol"].upper(), h["exchange"].upper())
        if key not in merged:
            merged[key] = {"symbol": h["symbol"], "exchange": h["exchange"], "qty": 0.0, "cost": 0.0}
            order.append(key)
        merged[key]["qty"] += h["qty"]
        merged[key]["cost"] += h["qty"] * h["buy_price"]

    result = []
    for key in order:
        m = merged[key]
        avg_buy_price = m["cost"] / m["qty"] if m["qty"] else 0.0
        result.append({
            "symbol": m["symbol"],
            "exchange": m["exchange"],
            "qty": m["qty"],
            "buy_price": round(avg_buy_price, 4),
        })
    return result


def fetch_quote(symbol, exchange):
    """Returns (ltp, prev_close) for a symbol.
    prev_close is the prior session's close, used to compute today's change."""
    ticker = symbol.strip().upper() + EXCHANGE_SUFFIX[exchange.upper()]
    t = yf.Ticker(ticker)

    ltp = None
    prev_close = None

    # fast_info reflects the latest traded price during market hours,
    # more current than the daily history bar.
    try:
        fast = t.fast_info
        price = fast.get("last_price")
        if price:
            ltp = round(float(price), 2)
        prev = fast.get("previous_close")
        if prev:
            prev_close = round(float(prev), 2)
    except Exception:
        pass

    if ltp is None or prev_close is None:
        data = t.history(period="5d")
        if data.empty:
            raise ValueError(f"No price data for {ticker}")

        if ltp is None:
            # Most recent 1-minute bar (still current during market hours)
            intraday = t.history(period="1d", interval="1m")
            if not intraday.empty:
                ltp = round(float(intraday["Close"].iloc[-1]), 2)
            else:
                # Market closed (weekend/holiday) - fall back to last daily close
                ltp = round(float(data["Close"].iloc[-1]), 2)

        if prev_close is None:
            if len(data) >= 2:
                prev_close = round(float(data["Close"].iloc[-2]), 2)
            else:
                prev_close = ltp  # no prior session available; treat day change as 0

    return ltp, prev_close


def _fetch_quote_safe(symbol, exchange):
    """fetch_quote wrapped for use with ThreadPoolExecutor.map, which can't
    propagate a per-item exception without aborting the whole batch."""
    try:
        ltp, prev_close = fetch_quote(symbol, exchange)
        return ltp, prev_close, None
    except Exception as e:
        return None, None, str(e)


def compute_rows(holdings):
    """Fetch prices and compute P&L for each holding. Returns rows + totals.
    Price fetches are independent network calls, so they run concurrently
    instead of one after another."""
    rows = []
    total_invested = 0.0
    total_current = 0.0
    total_prev_value = 0.0
    total_day_pnl = 0.0

    with ThreadPoolExecutor(max_workers=min(8, len(holdings) or 1)) as executor:
        quotes = list(executor.map(lambda h: _fetch_quote_safe(h["symbol"], h["exchange"]), holdings))

    for h, (ltp, prev_close, error) in zip(holdings, quotes):
        if error is not None:
            rows.append({"symbol": h["symbol"], "error": error})
            continue

        invested = h["qty"] * h["buy_price"]
        current = h["qty"] * ltp
        pnl = current - invested
        pnl_pct = (pnl / invested) * 100 if invested else 0

        prev_value = h["qty"] * prev_close
        day_pnl = current - prev_value
        day_pct = ((ltp - prev_close) / prev_close) * 100 if prev_close else 0

        total_invested += invested
        total_current += current
        total_prev_value += prev_value
        total_day_pnl += day_pnl

        rows.append({
            "symbol": h["symbol"],
            "exchange": h["exchange"],
            "qty": h["qty"],
            "buy_price": h["buy_price"],
            "ltp": ltp,
            "prev_close": prev_close,
            "invested": invested,
            "current": current,
            "pnl": pnl,
            "pnl_pct": pnl_pct,
            "day_pnl": day_pnl,
            "day_pct": day_pct,
        })

    total_pnl = total_current - total_invested
    total_pct = (total_pnl / total_invested) * 100 if total_invested else 0
    total_day_pct = (total_day_pnl / total_prev_value) * 100 if total_prev_value else 0
    totals = {
        "invested": total_invested,
        "current": total_current,
        "pnl": total_pnl,
        "pnl_pct": total_pct,
        "day_pnl": total_day_pnl,
        "day_pct": total_day_pct,
    }
    return rows, totals


def build_telegram_message(rows, totals):
    """Short headline summary for the chat message; the full per-stock
    breakdown goes out separately as a PDF attachment (see build_pdf_report)."""
    timestamp = now_ist().strftime("%d %b, %I:%M %p IST")

    total_arrow = "🟢" if totals["pnl"] >= 0 else "🔴"
    total_day_arrow = "🟢" if totals["day_pnl"] >= 0 else "🔴"
    failed = [r["symbol"] for r in rows if "error" in r]

    lines = [
        f"📊 *Portfolio P&L* — _{timestamp}_",
        "",
        f"{total_arrow} *Total P&L: ₹{totals['pnl']:+,.2f} ({totals['pnl_pct']:+.2f}%)*",
        f"{total_day_arrow} *Today: ₹{totals['day_pnl']:+,.2f} ({totals['day_pct']:+.2f}%)*",
        f"Investment: ₹{totals['invested']:,.2f}  •  Value: ₹{totals['current']:,.2f}",
        "",
        "📄 Full stock-wise breakdown attached as PDF.",
    ]
    if failed:
        lines.append(f"⚠️ Price fetch failed for: {', '.join(failed)}")

    return "\n".join(lines)


# Shared column definitions so the email table and the PDF table stay
# in sync — same fields, same order, same green/red coloring rules.
REPORT_COLUMNS = [
    ("symbol", "Symbol", "left", None),
    ("exchange", "Exch", "left", None),
    ("qty", "Qty", "right", None),
    ("buy_price", "Buy Price", "right", None),
    ("ltp", "LTP", "right", None),
    ("invested", "Total Investment", "right", None),
    ("current", "Market Value", "right", None),
    ("pnl", "P&L", "right", "pnl"),
    ("pnl_pct", "P&L %", "right", "pnl"),
    ("day_pnl", "Today's P&L", "right", "day"),
    ("day_pct", "Today's %", "right", "day"),
]


def format_report_rows(rows, totals):
    """Formats holdings + totals into the shared cell text/coloring used by
    both the email table and the PDF table. Returns (body, total_row) where
    body entries are either {"error": True, "symbol": ...} or
    {"error": False, "cells": {col_key: text}, "pnl_positive": bool, "day_positive": bool}."""
    body = []
    for r in rows:
        if "error" in r:
            body.append({"error": True, "symbol": r["symbol"]})
            continue
        cells = {
            "symbol": r["symbol"],
            "exchange": r["exchange"],
            "qty": f"{r['qty']:g}",
            "buy_price": f"{r['buy_price']:,.2f}",
            "ltp": f"{r['ltp']:,.2f}",
            "invested": f"{r['invested']:,.2f}",
            "current": f"{r['current']:,.2f}",
            "pnl": f"{r['pnl']:+,.2f}",
            "pnl_pct": f"{r['pnl_pct']:+.2f}%",
            "day_pnl": f"{r['day_pnl']:+,.2f}",
            "day_pct": f"{r['day_pct']:+.2f}%",
        }
        body.append({"error": False, "cells": cells, "pnl_positive": r["pnl"] >= 0, "day_positive": r["day_pnl"] >= 0})

    total_row = {
        "cells": {
            "symbol": "TOTAL", "exchange": "", "qty": "", "buy_price": "", "ltp": "",
            "invested": f"{totals['invested']:,.2f}",
            "current": f"{totals['current']:,.2f}",
            "pnl": f"{totals['pnl']:+,.2f}",
            "pnl_pct": f"{totals['pnl_pct']:+.2f}%",
            "day_pnl": f"{totals['day_pnl']:+,.2f}",
            "day_pct": f"{totals['day_pct']:+.2f}%",
        },
        "pnl_positive": totals["pnl"] >= 0,
        "day_positive": totals["day_pnl"] >= 0,
    }
    return body, total_row


def build_pdf_report(rows, totals):
    """Builds a formatted PDF with the full per-stock P&L table, for
    a properly rendered table regardless of the viewer's device/app."""
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        topMargin=12 * mm, bottomMargin=12 * mm, leftMargin=10 * mm, rightMargin=10 * mm,
    )
    styles = getSampleStyleSheet()
    timestamp = now_ist().strftime("%d %b %Y, %I:%M %p IST")

    green = colors.HexColor("#1a7f37")
    red = colors.HexColor("#cf222e")

    body, total_row = format_report_rows(rows, totals)
    n_cols = len(REPORT_COLUMNS)

    data = [[label for _, label, _, _ in REPORT_COLUMNS]]
    cell_colors = []  # (row, col, color)

    for i, r in enumerate(body, start=1):
        if r["error"]:
            data.append([r["symbol"]] + [""] * (n_cols - 2) + ["price fetch failed"])
            continue
        data.append([r["cells"][key] for key, _, _, _ in REPORT_COLUMNS])
        for j, (key, _, _, group) in enumerate(REPORT_COLUMNS):
            if group == "pnl":
                cell_colors.append((i, j, green if r["pnl_positive"] else red))
            elif group == "day":
                cell_colors.append((i, j, green if r["day_positive"] else red))

    total_idx = len(data)
    data.append([total_row["cells"][key] for key, _, _, _ in REPORT_COLUMNS])
    for j, (key, _, _, group) in enumerate(REPORT_COLUMNS):
        if group == "pnl":
            cell_colors.append((total_idx, j, green if total_row["pnl_positive"] else red))
        elif group == "day":
            cell_colors.append((total_idx, j, green if total_row["day_positive"] else red))

    table = Table(data, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a2e")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, total_idx), (-1, total_idx), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dddddd")),
        ("ROWBACKGROUNDS", (0, 1), (-1, total_idx - 1), [colors.white, colors.HexColor("#f7f7fb")]),
        ("BACKGROUND", (0, total_idx), (-1, total_idx), colors.HexColor("#f0f0f5")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]
    for j, (_, _, align, _) in enumerate(REPORT_COLUMNS):
        style.append(("ALIGN", (j, 0), (j, -1), align.upper()))
    for row, col, color in cell_colors:
        style.append(("TEXTCOLOR", (col, row), (col, row), color))
    table.setStyle(TableStyle(style))

    elements = [
        Paragraph(f"<b>Portfolio P&amp;L — {timestamp}</b>", styles["Title"]),
        Spacer(1, 8),
        table,
    ]
    doc.build(elements)
    return buf.getvalue()


def build_email_html(rows, totals):
    timestamp = now_ist().strftime("%d %b %Y, %I:%M %p IST")

    body, total_row = format_report_rows(rows, totals)
    n_cols = len(REPORT_COLUMNS)

    def cell_color(group, positive):
        if group is None:
            return ""
        return "#1a7f37" if positive else "#cf222e"

    row_html = ""
    for r in body:
        if r["error"]:
            row_html += (
                f"<tr><td style='padding:8px;border-bottom:1px solid #eee;'>{r['symbol']}</td>"
                f"<td colspan='{n_cols - 1}' style='padding:8px;border-bottom:1px solid #eee;color:#999;'>"
                f"price fetch failed</td></tr>"
            )
            continue
        bg = "#f0fdf4" if r["pnl_positive"] else "#fef2f2"
        cells_html = ""
        for key, _, align, group in REPORT_COLUMNS:
            positive = r["pnl_positive"] if group == "pnl" else r["day_positive"]
            color = cell_color(group, positive)
            style = f"padding:8px;border-bottom:1px solid #eee;text-align:{align};"
            if color:
                style += f"color:{color};font-weight:600;"
            cells_html += f'<td style="{style}">{r["cells"][key]}</td>'
        row_html += f'<tr style="background:{bg};">{cells_html}</tr>'

    header_html = "".join(
        f'<th style="padding:8px;text-align:{align};">{label}</th>' for _, label, align, _ in REPORT_COLUMNS
    )
    total_cells_html = ""
    for key, _, align, group in REPORT_COLUMNS:
        positive = total_row["pnl_positive"] if group == "pnl" else total_row["day_positive"]
        color = cell_color(group, positive)
        style = f"padding:10px 8px;text-align:{align};"
        if color:
            style += f"color:{color};"
        total_cells_html += f'<td style="{style}">{total_row["cells"][key]}</td>'

    html = f"""
    <html><body style="font-family:Segoe UI,Arial,sans-serif;background:#f6f6f6;padding:20px;">
      <div style="max-width:820px;margin:auto;background:#fff;border-radius:8px;overflow:hidden;box-shadow:0 1px 4px rgba(0,0,0,0.1);">
        <div style="background:#1a1a2e;color:#fff;padding:16px 20px;">
          <h2 style="margin:0;font-size:18px;">📊 Portfolio P&amp;L</h2>
          <div style="font-size:12px;color:#aaa;margin-top:4px;">{timestamp}</div>
        </div>
        <table style="width:100%;border-collapse:collapse;font-size:14px;">
          <thead>
            <tr style="background:#f0f0f5;">{header_html}</tr>
          </thead>
          <tbody>
            {row_html}
            <tr style="background:#f0f0f5;font-weight:700;">{total_cells_html}</tr>
          </tbody>
        </table>
      </div>
    </body></html>
    """
    return html


def send_telegram(text):
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
    resp = requests.post(url, data={"chat_id": CHAT_ID, "text": text, "parse_mode": "Markdown"})
    resp.raise_for_status()


def send_telegram_document(file_bytes, filename, caption=None):
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument"
    data = {"chat_id": CHAT_ID}
    if caption:
        data["caption"] = caption
    files = {"document": (filename, file_bytes, "application/pdf")}
    resp = requests.post(url, data=data, files=files)
    resp.raise_for_status()


def send_email(html_body):
    subject = f"Portfolio P&L — {now_ist().strftime('%d %b, %I:%M %p IST')}"
    url = "https://api.brevo.com/v3/smtp/email"
    headers = {
        "api-key": BREVO_API_KEY,
        "Content-Type": "application/json",
        "Accept": "application/json",
    }
    payload = {
        "sender": {"email": EMAIL_FROM},
        "to": [{"email": EMAIL_TO}],
        "subject": subject,
        "htmlContent": html_body,
    }
    resp = requests.post(url, json=payload, headers=headers)
    resp.raise_for_status()


def _check_required_env_vars():
    missing = [name for name, val in (("BOT_TOKEN", BOT_TOKEN), ("CHAT_ID", CHAT_ID), ("BREVO_API_KEY", BREVO_API_KEY), ("EMAIL_FROM", EMAIL_FROM), ("EMAIL_TO", EMAIL_TO)) if not val]
    if missing:
        print("Missing required environment variables:", ", ".join(missing), file=sys.stderr)
        print("Make sure GitHub Actions secrets (or local environment variables) are set.", file=sys.stderr)
        sys.exit(1)


def main():
    # Fail early with a clear message if secrets are missing
    _check_required_env_vars()

    holdings = load_portfolio(PORTFOLIO_FILE)
    rows, totals = compute_rows(holdings)

    telegram_text = build_telegram_message(rows, totals)
    email_html = build_email_html(rows, totals)
    pdf_bytes = build_pdf_report(rows, totals)

    print(telegram_text)  # for local testing visibility

    try:
        send_telegram(telegram_text)
        print("✅ Telegram sent.")
    except Exception as e:
        print(f"⚠️ Telegram failed: {e}", file=sys.stderr)

    try:
        pdf_filename = f"portfolio_pnl_{now_ist().strftime('%Y%m%d_%H%M')}.pdf"
        send_telegram_document(pdf_bytes, pdf_filename)
        print("✅ Telegram PDF sent.")
    except Exception as e:
        print(f"⚠️ Telegram PDF failed: {e}", file=sys.stderr)

    try:
        send_email(email_html)
        print("✅ Email sent.")
    except Exception as e:
        print(f"⚠️ Email failed: {e}", file=sys.stderr)


if __name__ == "__main__":
    main()
